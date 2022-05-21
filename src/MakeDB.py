import pymysql.cursors
import openpyxl
import pykakasi


# Connect to the database
def MakeFood():
    wb = openpyxl.load_workbook("1365334_1r12.xlsx")
    ws = wb.worksheets[0]
    kks = pykakasi.kakasi()
    for Row in range(9, 2200):
        FoodName = ws.cell(row=Row, column=4).value
        result = kks.convert(FoodName)
        FoodName_kana = ''
        for char in result:
            FoodName_kana += char['kana']
        
        try:
            Kcal = float(ws.cell(row=Row, column=6).value)
        except ValueError:
            Kcal = float(0)
        try:
            Protein = float(ws.cell(row=Row, column=9).value)
        except ValueError:
            protein = float(0)
        try:
            Lipids = float(ws.cell(row=Row, column=11).value)
        except ValueError:
            Lipids = float(0)
        try:
            Carbohydrate =  float(ws.cell(row=Row, column=17).value)
        except ValueError:
            Carbohydrate = float(0)
        MakeFoodDB(FoodName, FoodName_kana, Kcal, Protein, Lipids, Carbohydrate)
        
            


def MakeFoodDB(FoodName, FoodName_kana, Kcal, Protein, Lipids, Carbohydrate):
    # FoodName = 名前, FoodName = 名前(カナ) Kcal = カロリー, Protein = タンパク質, Lipids = 脂質, Carbohyrate = 炭水化物)
    connection = pymysql.connect(host = '127.0.0.1',
                                user = 'root',
                                password = '',
                                db = 'Hackathon',
                                charset = 'utf8mb4',
                                cursorclass=pymysql.cursors.DictCursor)
    try:
        connection.begin()
        with connection.cursor() as cursor:
            cursor = connection.cursor()
            # SQLの作成、定義
            sql = "insert into Food(FoodName, FoodName_kana, Kcal, Protein, Lipids, Carbohydrate) value(%s, %s, %s, %s, %s, %s)"
            # SQLの実行
            cursor.execute(sql, (FoodName, FoodName_kana, Kcal, Protein, Lipids, Carbohydrate))
            results = cursor.fetchall()
            for i, row in enumerate(results):
                print(i, row)

            connection.commit()

    except Exception as e:
      print('error:', e)
      connection.rollback()
    finally:
      connection.close()



def Main():
    tmp = MakeFood()

if __name__ == '__main__':
    Main()
